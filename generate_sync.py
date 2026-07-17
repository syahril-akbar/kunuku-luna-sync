import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
import difflib
import re
import os
import sys
import shutil
import warnings

# Bungkam peringatan bawel dari openpyxl (DrawingML warning)
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

OUTLET_CATEGORIES = {
    "PRT": "PERINTIS",
    "HRT": "HERTASNING",
    "MLG": "MALANG"
}

# ============================================================
# KONFIGURASI SUB KATEGORI (Single Source of Truth)
# Tambahkan sub kategori baru di sini untuk memperluas sistem.
# ============================================================
# MODE yang tersedia:
#   'prefix'       -> Selalu tambahkan 'abbr' di depan nama produk.
#                     Contoh: BUTTER RICE -> 'BR Chicken Rice Mentai'
#   'keyword'      -> Tambahkan 'abbr' HANYA jika nama belum mengandung
#                     salah satu 'keywords'. Mencegah duplikasi.
#                     Contoh: SOUP -> 'SUP Empal Hati Sapi'
#                              tapi 'Sop Merah Surabaya' dibiarkan (sudah ada 'SOP')
#   'strip_prefix' -> Hapus awalan sub kategori ('strip') dari nama jika ada,
#                     lalu tambahkan 'abbr'. Mencegah redundansi nama.
#                     Contoh: FINGER FOOD -> 'FF Mie Original' (bukan 'FF Finger Food Mie Original')
#   'pass'         -> Tidak ada perubahan. Nama produk sudah self-descriptive
#                     (sudah mengandung identitas kategori atau tidak perlu prefix).
# ============================================================
SUB_KATEGORI_CONFIG = {
    # Sub kategori yang perlu singkatan/prefix:
    'SOUP':       {'mode': 'keyword',      'abbr': 'SUP', 'keywords': r'\b(SUP|SOUP|SOP|SOTO)\b'},
    'BUTTER RICE':{'mode': 'strip_prefix', 'abbr': 'BR',  'strip': r'\bBUTTER\s+RICE\b\s*'},
    'FINGER FOOD':{'mode': 'strip_prefix', 'abbr': 'FF',  'strip': r'^FING+ER\s+FOOD\s*'},
    'RICE BOX':   {'mode': 'strip_prefix', 'abbr': 'RB',  'strip': r'^RICE\s+BOX\s*'},
    'SNACK BUAH': {'mode': 'strip_prefix', 'abbr': 'SB',  'strip': r'^SNACK\s+BUAH\s*'},  # singkatan SB untuk pengelompokan

    # Sub kategori bubur (dihandle oleh regex bubur sebelum cek sub kategori):
    'PUREE':      {'mode': 'pass'},
    'SEMI SOLID': {'mode': 'pass'},
    'SOLID':      {'mode': 'pass'},

    # Sub kategori dengan nama produk yang biasanya self-descriptive,
    # tapi pakai mode 'keyword' untuk jaga-jaga jika ada produk baru
    # yang namanya tidak diawali nama kategori:
    'KALDU': {'mode': 'keyword', 'abbr': 'KALDU', 'keywords': r'\bKALDU\b'},
    'LAUK':  {'mode': 'keyword', 'abbr': 'LAUK',  'keywords': r'\bLAUK\b'},
    'PASTA': {'mode': 'keyword', 'abbr': 'PASTA',  'keywords': r'\b(PASTA|MACARONI)\b'},
    'ABON':  {'mode': 'keyword', 'abbr': 'ABON',   'keywords': r'\bABON\b'},
    'GHEE & SAUS': {'mode': 'keyword', 'abbr': 'GHEE', 'keywords': r'\b(GHEE|SAUS)\b'},
    'PELENGKAP BB BOOSTER':{'mode': 'prefix', 'abbr': 'BB'},  # singkatan BB untuk pengelompokan
    'CUP, STIKER, & PLASTIK': {'mode': 'pass'},  # non-makanan, tidak perlu prefix
    'MAINAN/AKSESORIS':       {'mode': 'pass'},  # dihandle terpisah oleh logika mainan_keywords di format_nama_luna()
}

def auto_resize_columns(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Dapatkan huruf kolom
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 60)

def apply_color_grouping_to_sheet(ws):
    header_row_idx = 3
    data_rows = []
    
    # Baca semua baris data
    for r in range(header_row_idx + 1, ws.max_row + 1):
        row_values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if not row_values[2]:
            continue
        data_rows.append(row_values)
        
    if not data_rows:
        return
        
    # Tentukan grup kategori untuk setiap baris data
    for row in data_rows:
        prod_name = row[2]
        # Get prefix
        name = str(prod_name).strip()
        match_prefix = re.match(r'^([A-Z]{3})\s+(.*)$', name, re.IGNORECASE)
        if match_prefix:
            name = match_prefix.group(2).strip()
        
        name_upper = name.upper()
        
        # BUBUR dengan ukuran ML
        if "BUBUR" in name_upper:
            ml_match = re.search(r'\b(\d+)\s*ML\b', name_upper)
            if ml_match:
                group = f"BUBUR {ml_match.group(1)} ML"
            else:
                group = "BUBUR"
        else:
            # Cek mainan
            mainan_keywords = ['LONCENG', 'GENDERANG', 'KENYOT', 'GELANG', 'TEPUK TANGAN', 'TEETER', 'DOT BABY', 'MAINAN']
            if any(k in name_upper for k in mainan_keywords):
                group = "MAINAN/AKSESORIS"
            else:
                # Cek sub kategori config
                found = False
                for key, cfg in SUB_KATEGORI_CONFIG.items():
                    abbr = cfg.get('abbr')
                    if abbr and (name_upper.startswith(abbr.upper() + " ") or name_upper.startswith(abbr.upper() + ".")):
                        group = key
                        found = True
                        break
                    if key in name_upper:
                        group = key
                        found = True
                        break
                if not found:
                    words = name.split()
                    group = words[0].upper() if words else "LAIN-LAIN"
        row.append(group)
        
    # Sort data_rows berdasarkan group, lalu nama produk
    data_rows.sort(key=lambda x: (x[-1], x[2]))
    
    # Hapus baris data lama dari sheet
    if ws.max_row > header_row_idx:
        ws.delete_rows(header_row_idx + 1, ws.max_row - header_row_idx)
        
    pastel_colors = [
        "ADD8E6", # Light Blue
        "90EE90", # Light Green
        "FFF3CD", # Light Yellow/Gold
        "FFB6C1", # Light Pink
        "E6E6FA", # Lavender/Purple
        "FFA07A", # Light Salmon/Orange
        "87CEFA", # Light Sky Blue
        "F0E68C", # Khaki
        "20B2AA", # Light Sea Green/Teal
        "D3D3D3"  # Light Grey
    ]
    
    group_colors = {}
    color_idx = 0
    
    # Tulis kembali data yang sudah di-sort dan diwarnai
    for i, row in enumerate(data_rows, 1):
        group = row[-1]
        row_data = row[:-1]
        row_data[0] = i # Re-numbering No
        
        ws.append(row_data)
        curr_row = ws.max_row
        
        if group not in group_colors:
            group_colors[group] = pastel_colors[color_idx % len(pastel_colors)]
            color_idx += 1
            
        color_hex = group_colors[group]
        fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        
        for c in range(1, len(row_data) + 1):
            cell = ws.cell(curr_row, c)
            cell.fill = fill


def safe_int(value):
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    digits = re.sub(r'[^\d]', '', str(value))
    return int(digits) if digits else 0

def safe_save_excel(wb, filepath):
    while True:
        try:
            wb.save(filepath)
            break
        except PermissionError:
            print(f"\n❌ ERROR: File '{os.path.basename(filepath)}' sedang dibuka oleh program lain (seperti Excel).")
            input("Silakan tutup file tersebut di Excel, lalu tekan ENTER untuk mencoba menyimpan kembali...")

print("Memulai proses sinkronisasi dan pencocokan data...")

# 1. Validasi File Surat Jalan
while True:
    sj_filename = input("\nMasukkan nama/path file Surat Jalan (.xlsx): ").strip(" '\"")
    if not sj_filename.endswith('.xlsx'):
        sj_filename += '.xlsx'
    
    if os.path.exists(sj_filename):
        break
    else:
        print(f"ERROR: File '{sj_filename}' tidak ditemukan!")
        print("File yang tersedia di folder ini:")
        xlsx_files = [f for f in os.listdir('.') if f.endswith('.xlsx')]
        for f in xlsx_files:
            print(f"  - {f}")

# 2. Pemilihan Cabang (Strict Selection)
print("\nPilih Cabang Target:")
options = list(OUTLET_CATEGORIES.keys())
for idx, code in enumerate(options, 1):
    print(f"{idx}. {code} ({OUTLET_CATEGORIES[code]})")

while True:
    choice = input(f"Pilih nomor (1-{len(options)}): ").strip()
    if choice.isdigit() and 1 <= int(choice) <= len(options):
        warehouse_prefix = options[int(choice) - 1]
        break
    print("Input tidak valid! Pilih nomor yang tersedia.")

print(f"[OK] Menggunakan awalan: {warehouse_prefix} ({OUTLET_CATEGORIES[warehouse_prefix]})")

base_name = os.path.splitext(os.path.basename(sj_filename))[0]
suffix_text = re.sub(r'(?i)SURAT JALAN', '', base_name)
suffix_text = re.sub(r'(?i)FIX', '', suffix_text)
file_suffix = re.sub(r'[^A-Za-z0-9]+', '_', suffix_text).strip('_')
if not file_suffix:
    file_suffix = "OUTPUT"

# 3. Setup Folder Output (Berdasarkan Nama Cabang + Tanggal dr Nama File)
branch_name = OUTLET_CATEGORIES[warehouse_prefix]

# BERSIHKAN file_suffix dari redundancy nama cabang
# Hapus prefix (misal 'PRT') dan nama cabang (misal 'PERINTIS') dari suffix_text
clean_suffix = file_suffix
to_remove = [warehouse_prefix, branch_name, "OUTLET", "OUTET"] # Tambahkan kata umum yang sering muncul
for word in to_remove:
    # Gunakan RegEx untuk menghapus kata secara case-insensitive
    clean_suffix = re.sub(rf'(?i){word}', '', clean_suffix).strip('_')

# Pastikan tidak ada double underscore atau underscore menggantung
clean_suffix = re.sub(r'_{2,}', '_', clean_suffix).strip('_')

if not clean_suffix:
    clean_suffix = "OUTPUT"

folder_name = f"{branch_name}_{clean_suffix}"
output_dir = os.path.join(os.getcwd(), folder_name)

if not os.path.exists(output_dir):
    os.makedirs(output_dir)
    print(f"[NEW] Membuat folder baru: {folder_name}")
else:
    print(f"[DIR] Menggunakan folder yang ada: {folder_name}")

try:
    if not os.path.exists('Produk.xlsx'):
        print("❌ ERROR: File 'Produk.xlsx' (Master Data Luna) tidak ditemukan di folder root!")
        sys.exit(1)
        
    wb_produk = openpyxl.load_workbook('Produk.xlsx', data_only=True)
    ws_produk = wb_produk.active # Gunakan sheet aktif
    
    luna_data = {}
    all_luna_skus = set()
    for r in range(2, ws_produk.max_row + 1):
        sku = ws_produk.cell(r, 1).value
        nama = str(ws_produk.cell(r, 2).value or "").strip()
        qty = ws_produk.cell(r, 4).value
        harga_jual = ws_produk.cell(r, 6).value
        harga_modal = ws_produk.cell(r, 8).value
        
        if sku and nama and str(sku).strip().lower() != 'perintis' and nama.lower() != 'none':
            sku_str = str(sku).strip()
            all_luna_skus.add(sku_str)
            
            nama_str = nama.strip()
            # FILTER: Blokir SKU cabang lain (misal MLG) menyusup ke data PRT/HRT
            if not nama_str.upper().startswith(warehouse_prefix):
                continue
                
            luna_data[sku_str] = {
                'nama': nama_str,
                'qty': qty,
                'harga_jual': harga_jual,
                'harga_modal': harga_modal
            }
except Exception as e:
    print(f"❌ ERROR: Gagal membaca 'Produk.xlsx'. Pastikan file tidak sedang dibuka. ({e})")
    sys.exit(1)

# --- DETEKSI SHEET TERBAIK ---
try:
    wb_sj = openpyxl.load_workbook(sj_filename, data_only=True)
    ws_sj = wb_sj.active
    biggest_data_count = 0
    for sheet in wb_sj.sheetnames:
        temp_ws = wb_sj[sheet]
        # Hitung baris yang beneran ada isinya (cek kolom A-C di 20 baris pertama)
        data_count = 0
        for r in range(1, min(temp_ws.max_row + 1, 50)):
            if any(temp_ws.cell(r, c).value for c in range(1, 4)):
                data_count += 1
        
        if data_count > biggest_data_count:
            biggest_data_count = data_count
            ws_sj = temp_ws
            
    print(f"--> Target Data: '{ws_sj.title}' (Estimasi {biggest_data_count} Baris Data)")
except Exception as e:
    print(f"❌ ERROR: Gagal membaca Surat Jalan. ({e})")
    sys.exit(1)

# --- SMART HEADER DETECTION (Cek 5 Baris Pertama) ---
header_map = {}
header_row_index = 1

for r in range(1, 6):
    temp_map = {}
    found_cols = 0
    for c in range(1, ws_sj.max_column + 1):
        val = str(ws_sj.cell(r, c).value or "").strip().upper()
        if val and val not in temp_map: # FIX: Jangan overwrite jika sudah ada (Ambil yang paling kiri)
            temp_map[val] = c
            found_cols += 1
    
    # Jika baris ini punya setidaknya 3 kolom utama, kita anggap ini header
    main_cols = ["NAMA", "QTY", "ISI", "ITEM", "HARGA", "QUANTITY", "SKU", "ID"]
    match_count = sum(1 for k in temp_map.keys() if any(m in k for m in main_cols))
    
    if match_count >= 3:
        header_map = temp_map
        header_row_index = r
        break

if not header_map:
    print("ERROR: Gagal mendeteksi header di file Surat Jalan!")
    print("Pastikan ada kolom dengan nama: ID PRODUK, NAMA BARANG, QTY, dll.")
    sys.exit(1)

print(f"--> Header terdeteksi di Baris {header_row_index}")

# Map indices with aliases
def get_col_index(aliases, default):
    for alias in aliases:
        if alias.upper() in header_map:
            return header_map[alias.upper()]
    return default

col_id = get_col_index(["ID PRODUK", "SKU", "ID"], 1)
col_nama = get_col_index(["NAMA PRODUK", "NAMA BARANG", "NAMA"], 2)
# Prioritaskan 'TOTAL QTY'/'TOTAL PO' jika ada, kalau cuma ada 'TOTAL' (seperti di Perintis) baru ambil 'TOTAL'
col_qty = get_col_index(["TOTAL QTY", "TOTAL PO", "QTY", "QUANTITY", "TOTAL"], 5)
col_sat = get_col_index(["SATUAN", "SAT"], 6)
# Prioritaskan 'HARGA SATUAN'/'UNIT PRICE' untuk menghindari kerancuan dengan total harga (Total)
col_harga = get_col_index(["HARGA SATUAN", "HARGA", "UNIT PRICE"], 9)
col_subkat = get_col_index(["SUB KATEGORI", "SUBKATEGORI"], None)

print(f"--> Column Mapping: ID({col_id}), Nama({col_nama}), Qty({col_qty}), Sat({col_sat}), Harga({col_harga}), SubKat({col_subkat})")

# --- PHASE 1: SJ DATA COLLECTION & GROUND TRUTH AUDIT ---
sj_data = {}
baris_gagal = 0
baris_dilewati_bk = 0

# Variabel Auditing (Ground Truth)
target_total_qty = 0
target_total_rp = 0
target_total_items = 0
total_merges = 0
merged_details = []
names_to_ids = {} # Untuk deteksi konflik nama
name_conflicts = []

for r in range(header_row_index + 1, ws_sj.max_row + 1):
    id_raw = ws_sj.cell(r, col_id).value
    nama_raw = ws_sj.cell(r, col_nama).value
    nama = str(nama_raw or "").strip()
    qty = safe_int(ws_sj.cell(r, col_qty).value)
    sat = str(ws_sj.cell(r, col_sat).value or "").strip()
    harga = safe_int(ws_sj.cell(r, col_harga).value)
    subkat = str(ws_sj.cell(r, col_subkat).value or "").strip() if col_subkat else ""
    
    # Filter baris sampah/kosong
    if not nama or nama.lower() in ['none', 'nan', '']:
        continue
        
    # Kecualikan SKU yang berawalan BK
    if id_raw and str(id_raw).strip().upper().startswith('BK'):
        baris_dilewati_bk += 1
        continue
        
    # Audit: Tambahkan ke Target Ground Truth
    target_total_qty += qty
    target_total_rp += (qty * harga)
    target_total_items += 1

    if id_raw:
        id_str = str(id_raw).strip()
        
        # Deteksi Konflik Nama (Nama sama tapi ID beda)
        name_norm = nama.upper()
        if name_norm in names_to_ids:
            if id_str not in names_to_ids[name_norm]:
                names_to_ids[name_norm].append(id_str)
                if name_norm not in name_conflicts:
                    name_conflicts.append(name_norm)
        else:
            names_to_ids[name_norm] = [id_str]

        if id_str in sj_data:
            # DETEKSI DUPLIKASI / MERGE (ID SAMA)
            total_merges += 1
            merged_details.append(f"{id_str} ({nama})")
            sj_data[id_str]['qty'] += qty
        else:
            sj_data[id_str] = {
                'nama': nama,
                'qty': qty,
                'satuan': sat,
                'harga': harga,
                'sub_kategori': subkat
            }
    else:
        # Masuk ke kategori error jika nama ada tapi ID hilang
        baris_gagal += 1
        # Kita tetap proses sebagai barang tanpa ID untuk audit nanti
        temp_id = f"NO_ID_{r}"
        sj_data[temp_id] = {
            'nama': nama,
            'qty': qty,
            'satuan': sat,
            'harga': harga,
            'sub_kategori': subkat
        }

print(f"--> AUDIT INPUT: Terdeteksi {target_total_items} baris produk, Total Qty: {target_total_qty}, Total Rp: {target_total_rp:,}")

def normalize_spaces(text):
    return ' '.join(str(text).split()).upper()

def format_nama_luna(name, prefix, sub_kategori=''):
    name = normalize_spaces(name)
    name_upper = name.upper()
    sub_kat_upper = str(sub_kategori).strip().upper()
    
    # Koreksi typo Surat Jalan -> Standar Luna POS
    if 'GENDRANG' in name_upper:
        name = re.sub(r'(?i)GENDRANG', 'GENDERANG', name)
        name_upper = name.upper()
    
    # Deteksi Mainan dari keyword yang sudah kedaftar di Luna POS
    mainan_keywords = ['LONCENG', 'GENDERANG', 'KENYOT', 'GELANG', 'TEPUK TANGAN', 'TEETER', 'DOT BABY']
    is_mainan = any(k in name_upper for k in mainan_keywords)
    
    if is_mainan and 'SENDOK' not in name_upper:
        if 'MAINAN' not in name_upper:
            return f'{prefix} MAINAN {name}'
        else:
            if not name_upper.startswith(prefix):
                return f'{prefix} {name}'
            return name
            
    match_bubur = re.search(r'^(N[.,]?\s*TIM[.,]?|B\.+|B)\s+(.*?)\s+(\d+\+)\s*(\d+)\s*ML$', name)
    if match_bubur:
        varian = match_bubur.group(2).strip()
        umur = match_bubur.group(3).strip()
        volume = match_bubur.group(4).strip() + " ML"
        return f'{prefix} BUBUR {umur} {volume} {varian}'
    
    # Proteksi: Jika nama produk sudah berawalan BUBUR, lewati pemrosesan sub kategori
    if name_upper.startswith('BUBUR '):
        if not name_upper.startswith(prefix):
            return f'{prefix} {name}'
        return name
    
    # --- TERAPKAN ATURAN SUB KATEGORI dari SUB_KATEGORI_CONFIG ---
    if sub_kat_upper:
        cfg = SUB_KATEGORI_CONFIG.get(sub_kat_upper)
        if cfg is None:
            # Sub kategori baru/tidak dikenal: cetak warning agar bisa ditambahkan ke config
            print(f"  [PERINGATAN] Sub kategori tidak dikenal: '{sub_kategori}' (produk: {name}). "
                  f"Tambahkan ke SUB_KATEGORI_CONFIG di baris konfigurasi.")
        else:
            mode = cfg['mode']
            if mode == 'prefix':
                # Selalu tambahkan singkatan di depan
                name = f"{cfg['abbr']} {name}"
                name_upper = name.upper()
            elif mode == 'keyword':
                if sub_kat_upper == 'SOUP':
                    # Ganti SOP atau SOUP di awal nama dengan SUP
                    name = re.sub(r'^(SOP|SOUP)\b', 'SUP', name, flags=re.IGNORECASE).strip()
                    name_upper = name.upper()
                    # Selalu tambahkan SUP di awal jika belum ada
                    if not name_upper.startswith('SUP '):
                        name = f"SUP {name}"
                        name_upper = name.upper()
                else:
                    # Tambahkan singkatan hanya jika kata kunci belum ada di nama
                    if not re.search(cfg['keywords'], name_upper):
                        name = f"{cfg['abbr']} {name}"
                        name_upper = name.upper()
            elif mode == 'strip_prefix':
                # Hapus awalan sub kategori dari nama (cegah redundansi), lalu tambah singkatan
                candidate_name = re.sub(cfg['strip'], '', name, flags=re.IGNORECASE).strip()
                if sub_kat_upper == 'BUTTER RICE':
                    words_left = candidate_name.split()
                    if len(words_left) <= 1:
                        candidate_name = name
                name = candidate_name
                name_upper = name.upper()
                name = f"{cfg['abbr']} {name}"
                name_upper = name.upper()
            # mode == 'pass': tidak ada perubahan
    # --- END ATURAN SUB KATEGORI ---
    
    if name.startswith('KALDU'):
        return f'{prefix} {name}'
    if not name_upper.startswith(prefix):
        return f'{prefix} {name}'
    return name

matched_items_dict = {} # Key: SKU
unmatched_items_dict = {} # Key: Formatted Name
mapping_review = []
sku_collisions = [] # Untuk deteksi kalau ID SJ beda tapi SKU Luna sama

for id_sj, item_sj in sj_data.items():
    expected_luna_name = format_nama_luna(item_sj['nama'], warehouse_prefix, item_sj.get('sub_kategori', ''))
    
    best_match_sku = None
    best_match_nama_luna = None
    
    # 0. Prioritaskan Pencocokan ID (SKU) Mutlak terlebih dahulu
    # VALIDASI NAMA: Jika nama produk di Luna sangat berbeda dari nama SJ,
    # abaikan pencocokan ID (kemungkinan error data di Surat Jalan) dan fallback ke nama.
    if id_sj in luna_data:
        luna_nama_for_id = luna_data[id_sj]['nama']
        # Bandingkan kata-kata kunci dari nama SJ vs nama Luna (case-insensitive)
        # Ambil kata-kata bermakna (>= 3 huruf) dari masing-masing
        sj_words = set(w for w in re.findall(r'[A-Za-z]+', item_sj['nama'].upper()) if len(w) >= 3)
        luna_words = set(w for w in re.findall(r'[A-Za-z]+', luna_nama_for_id.upper()) if len(w) >= 3)
        # Hapus kata prefix outlet (HRT, MLG, PRT) dari luna_words
        luna_words.discard(warehouse_prefix)
        if sj_words and luna_words:
            common = sj_words & luna_words
            similarity = len(common) / max(len(sj_words), len(luna_words))
        else:
            similarity = 0.0
        
        if similarity >= 0.3:
            # Nama cukup mirip, pencocokan ID dianggap valid
            best_match_sku = id_sj
            best_match_nama_luna = luna_nama_for_id
        else:
            # Nama sangat berbeda -> kemungkinan ID salah di SJ, tandai sebagai KONFLIK
            print(f"  [KONFLIK ID] ID '{id_sj}' di SJ -> '{item_sj['nama']}' "
                  f"TIDAK COCOK dengan nama Luna: '{luna_nama_for_id}'. Fallback ke pencocokan nama.")
            # Simpan info konflik untuk ditampilkan di review
            item_sj['_id_konflik'] = f"ID {id_sj} = '{luna_nama_for_id}' di Luna"
    
    # 1. Exact Match by Nama (Fallback jika ID tidak cocok atau tidak ada)
    if not best_match_sku:
        expected_norm = normalize_spaces(expected_luna_name)
        for sku, data in luna_data.items():
            if normalize_spaces(data['nama']) == expected_norm:
                best_match_sku = sku
                best_match_nama_luna = data['nama']
                break
    
    # Tentukan status dan flag konflik ID
    id_konflik_info = item_sj.get('_id_konflik', '')
    mapping_review.append({
        'id_sj': id_sj,
        'nama_sj': item_sj['nama'],
        'sub_kategori': item_sj.get('sub_kategori', ''),
        'qty_sj': item_sj['qty'],
        'sku_luna_prediksi': best_match_sku,
        'nama_luna_prediksi': best_match_nama_luna if best_match_nama_luna else f"[NEW] {format_nama_luna(item_sj['nama'], warehouse_prefix, item_sj.get('sub_kategori', ''))}",
        'harga_satuan_sj': item_sj['harga'],
        'id_konflik': id_konflik_info
    })
    
    if best_match_sku:
        if best_match_sku in matched_items_dict:
            # TABRAKAN SKU (Beda ID SJ tapi 1 SKU LUNA)
            sku_collisions.append(f"{best_match_sku} ({best_match_nama_luna})")
            matched_items_dict[best_match_sku]['qty'] += item_sj['qty']
        else:
            matched_items_dict[best_match_sku] = {
                'sku': best_match_sku,
                'nama': best_match_nama_luna,
                'satuan': item_sj['satuan'],
                'qty': item_sj['qty']
            }
    else:
        new_name_key = format_nama_luna(item_sj['nama'], warehouse_prefix, item_sj.get('sub_kategori', ''))
        if new_name_key in unmatched_items_dict:
            unmatched_items_dict[new_name_key]['qty'] += item_sj['qty']
        else:
            unmatched_items_dict[new_name_key] = {
                'id': id_sj,
                'nama': item_sj['nama'],
                'sub_kategori': item_sj.get('sub_kategori', ''),
                'harga_satuan_sj': item_sj['harga'],
                'satuan': item_sj['satuan'],
                'qty': item_sj['qty']
            }

matched_items = list(matched_items_dict.values())
unmatched_items = list(unmatched_items_dict.values())

print(f"Selesai mapping. Ditemukan {len(matched_items)} prediksi sukses, {len(unmatched_items)} produk baru.")

wb_review = openpyxl.Workbook()
ws_review = wb_review.active
ws_review.title = "Review Mapping"

# URUTAN OPTIMAL (Mata Kiri ke Kanan): Status -> Sub Kategori -> Nama -> SKU -> Data
headers_review = ["STATUS MATCH", "Sub Kategori", "Nama dari Surat Jalan", "Nama POS LUNA", "SKU LUNA", "Qty Transfer", "Harga Satuan dr SJ", "SKU Surat Jalan"]
ws_review.append(headers_review)

# Style buat Header
for cell in ws_review[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    cell.alignment = Alignment(horizontal='center')

# Style buat Status
fill_new = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Kuning Muda (Baru)
fill_ok  = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Hijau Muda (OK)
fill_err = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid") # Merah Muda (Konflik)

for m in mapping_review:
    is_matched = True if m['sku_luna_prediksi'] else False
    id_konflik = m.get('id_konflik', '')
    
    if id_konflik:
        # Konflik ID: nama di SJ vs nama di Luna sangat berbeda
        status = f"KONFLIK ID (Cek Manual): {id_konflik}"
        fill_to_use = fill_err
    elif is_matched:
        status = "TERSEDIA DI POS"
        fill_to_use = fill_ok
    else:
        status = "PRODUK BARU"
        fill_to_use = fill_new
    
    row_data = [
        status,                            # Kolom 1: Status (Prioritas Mata)
        m.get('sub_kategori', ''),         # Kolom 2: Sub Kategori (untuk analisis/filter)
        m['nama_sj'],                      # Kolom 3: Nama Asli dari SJ
        m['nama_luna_prediksi'] or "N/A",  # Kolom 4: Nama di Luna POS
        m['sku_luna_prediksi'] or "N/A",   # Kolom 5: Kode SKU
        m['qty_sj'],                       # Kolom 6: Qty
        m['harga_satuan_sj'],              # Kolom 7: Harga
        m['id_sj']                         # Kolom 8: ID Referensi
    ]
    ws_review.append(row_data)
    
    # Beri warna pada baris yang baru saja ditambahkan
    curr_row = ws_review.max_row
    
    # Warnai kolom STATUS (Kolom 1)
    status_cell = ws_review.cell(row=curr_row, column=1)
    status_cell.fill = fill_to_use
    status_cell.font = Font(bold=True)
    status_cell.alignment = Alignment(horizontal='center')
    # Warnai kolom Sub Kategori (Kolom 2) dengan warna yang sama
    ws_review.cell(row=curr_row, column=2).fill = fill_to_use

auto_resize_columns(ws_review)
# Simpan ke folder output
review_file = os.path.join(output_dir, f"Laporan_Mapping_Review_{folder_name}.xlsx")
safe_save_excel(wb_review, review_file)

wb_tf = openpyxl.load_workbook('warehouse-transfer-import-template.xlsx')
ws_tf = wb_tf.active
if ws_tf.max_row >= 4:
    ws_tf.delete_rows(4, ws_tf.max_row - 3)

for i, m in enumerate(matched_items, 1):
    row = [i, m['sku'], m['nama'], m['satuan'], m['qty']] + ([None] * 20)
    ws_tf.append(row)
auto_resize_columns(ws_tf)
# Simpan ke folder output
tf_file = os.path.join(output_dir, f"Siap_Warehouse_Transfer_{folder_name}.xlsx")
safe_save_excel(wb_tf, tf_file)

wb_new = openpyxl.load_workbook('product-import-template.xlsx')
ws_new = wb_new.active
if ws_new.max_row >= 4:
    ws_new.delete_rows(4, ws_new.max_row - 3)

for i, u in enumerate(unmatched_items, 1):
    harga_jual_luna = u['harga_satuan_sj']
    nama_baru = format_nama_luna(u['nama'], warehouse_prefix, u.get('sub_kategori', ''))
    kategori = OUTLET_CATEGORIES.get(warehouse_prefix, warehouse_prefix)
    
    # Gunakan list row yang definitif
    row = [
        i,           # Col 1: No
        "",          # Col 2: SKU
        nama_baru,   # Col 3: Nama
        "Y",         # Col 4: Tersedia
        "N",         # Col 5: Bundle
        harga_jual_luna, # Col 6: Harga Jual
        0,           # Col 7: Harga Modal (DIPAKSA 0)
        "Y",         # Col 8: Stok Aktif
        u['qty'],    # Col 9: Jumlah Stok
        1,           # Col 10: Min Stok
        kategori,    # Col 11: Kategori
        str(u['satuan']).upper() if u['satuan'] else "PCS" # Col 12: Unit
    ] + ([None] * 5)
    
    ws_new.append(row)
    
    # Double check: Pastikan cell-nya beneran terisi angka 0 (Numeric)
    last_row = ws_new.max_row
    ws_new.cell(row=last_row, column=7).value = 0
apply_color_grouping_to_sheet(ws_new)
auto_resize_columns(ws_new)
# Simpan ke folder output
new_prod_file = os.path.join(output_dir, f"Siap_Product_Baru_{folder_name}.xlsx")
safe_save_excel(wb_new, new_prod_file)

# ---- FITUR LAPORAN OTOMATIS AKHIR ----
# ---- PHASE 4: FINAL RECONCILIATION AUDIT ----
total_qty_transfer = sum(m['qty'] for m in matched_items)
total_qty_baru = sum(u['qty'] for u in unmatched_items)
total_item_transfer = len(matched_items)
total_item_baru = len(unmatched_items)

total_rp_transfer = sum(m['qty_sj'] * m['harga_satuan_sj'] for m in mapping_review if m['sku_luna_prediksi'])
total_rp_baru = sum(m['qty_sj'] * m['harga_satuan_sj'] for m in mapping_review if not m['sku_luna_prediksi'])

# Hitung Akurasi
actual_total_qty = total_qty_transfer + total_qty_baru
actual_total_rp = total_rp_transfer + total_rp_baru
actual_total_items = total_item_transfer + total_item_baru

diff_qty = target_total_qty - actual_total_qty
diff_rp = target_total_rp - actual_total_rp

reconciliation_status = "PERFECT (MATCH)" if diff_qty == 0 and diff_rp == 0 else "ERROR (MISMATCH)"
reconciliation_color = "" # Bisa ditambah ANSI color jika dijalankan di terminal modern

rp_transfer_str = f"Rp {total_rp_transfer:,}".replace(',', '.')
rp_baru_str = f"Rp {total_rp_baru:,}".replace(',', '.')
diff_rp_str = f"Rp {diff_rp:,}".replace(',', '.')

laporan_text = f"""=========================================
LAPORAN SINKRONISASI INVENTORI LUNA POS
=========================================
File Sumber       : {os.path.basename(sj_filename)}
Cabang Target     : {warehouse_prefix}
Status Akurasi    : {reconciliation_status}

-- AUDIT REKONSILIASI (WAJIB NOL) --
Selisih Quantity  : {diff_qty} Pcs
Selisih Rp Value  : {diff_rp_str}
Total Baris SJ    : {target_total_items}
Total Baris Hasil : {actual_total_items}

-- AUDIT DUPLIKASI & KONFLIK --
ID Duplikat (Auto-Merge) : {total_merges} Item
Detail ID Terduplikasi   : {", ".join(merged_details) if merged_details else "Tidak ada"}
Tabrakan SKU Luna        : {len(sku_collisions)} Item
Detail SKU Tabrakan      : {", ".join(set(sku_collisions)) if sku_collisions else "Tidak ada"}
Konflik Nama di SJ       : {", ".join(name_conflicts) if name_conflicts else "Tidak ada (Aman)"}

-- RINCIAN PROSES --
Peringatan        : {baris_gagal} produk diproses tanpa ID/SKU asli (ID SJ Kosong)
Diabaikan         : {baris_dilewati_bk} produk di-skip karena SKU berawalan 'BK'

-- RINGKASAN MUTASI (STOCK IN LAMA) --
Total SKU Terdictasi               : {total_item_transfer} Varian
Total Quantitiy Masuk              : {total_qty_transfer} Pcs
Total Nilai Barang (Harga Satuan)  : {rp_transfer_str}

-- RINGKASAN REGISTRASI BARANG BARU --
Total SKU Baru Terdictasi          : {total_item_baru} Varian
Total Quantitiy Masuk              : {total_qty_baru} Pcs
Total Nilai Barang (Harga Satuan)  : {rp_baru_str}

-- DAFTAR FILE HASIL --
1. Laporan_Mapping_Review_{folder_name}.xlsx (WAJIB CEK)
2. Siap_Warehouse_Transfer_{folder_name}.xlsx
3. Siap_Product_Baru_{folder_name}.xlsx

========================================="""

laporan_filename = f"Laporan_Mutasi_{folder_name}.txt"
laporan_path = os.path.join(output_dir, laporan_filename)
with open(laporan_path, "w", encoding="utf-8") as f:
    f.write(laporan_text)

# Pindahkan file Surat Jalan (SJ) asli ke folder output untuk arsip
try:
    source_abs = os.path.abspath(sj_filename)
    target_sj_path = os.path.join(output_dir, os.path.basename(sj_filename))
    target_abs = os.path.abspath(target_sj_path)
    
    if source_abs != target_abs:
        if os.path.exists(target_abs):
            os.remove(target_abs)
        shutil.move(sj_filename, target_sj_path)
        print(f"[DONE] File sumber '{os.path.basename(sj_filename)}' dipindahkan ke folder {folder_name} untuk arsip.")
    else:
        print(f"[KEEP] File sumber '{os.path.basename(sj_filename)}' sudah berada di folder arsip.")
except PermissionError:
    print(f"PERINGATAN: File '{os.path.basename(sj_filename)}' sedang dibuka oleh program lain (Excel?).")
    print(f"   --> File GAGAL dipindahkan otomatis, tapi hasil generate DI DALAM FOLDER {folder_name} tetap aman.")
except Exception as e:
    print(f"Peringatan: Gagal memindahkan file sumber: {e}")

# Copy file Produk.xlsx ke folder output sebagai referensi
try:
    produk_src = 'Produk.xlsx'
    produk_dst = os.path.join(output_dir, 'Produk.xlsx')
    if os.path.exists(produk_src):
        shutil.copy2(produk_src, produk_dst)
        print(f"[COPY] File 'Produk.xlsx' di-copy ke folder {folder_name} sebagai referensi.")
    else:
        print(f"Peringatan: File 'Produk.xlsx' tidak ditemukan untuk di-copy.")
except Exception as e:
    print(f"Peringatan: Gagal meng-copy file Produk.xlsx: {e}")

print(f"\n{laporan_text}\n")
print(f"[SUCCESS] Semua file berhasil disimpan di folder: {output_dir}")
